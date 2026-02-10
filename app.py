# -*- coding: utf-8 -*-
import io
import os
import json
import sqlite3
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta

import bcrypt
import streamlit as st
from openpyxl import load_workbook

APP_TITLE = "安全指示KY（クラウド・ログイン版）"
DB_PATH = os.environ.get("KY_DB_PATH", "/tmp/ky_app.sqlite3")
TEMPLATE_PATH = os.environ.get("KY_TEMPLATE_PATH", "安全指示ＫＹ記録書.xlsx")
SEED_PATH = os.environ.get("KY_SEED_PATH", "seed.json")
RETENTION_YEARS = int(os.environ.get("KY_RETENTION_YEARS", "3"))

# ---- Excel cell mapping (based on the provided template) ----
# Top section (these feed the report section via formulas)
CELL = {
    "work_title": "C9",          # 作業件名 (merged C9:J9)
    "work_company": "C10",       # 作業会社名 (merged C10:D10)
    "phone": "G10",              # 電話番号 (merged G10:J10)
    "work_date": "C11",          # 作業予定日時 日付 (merged C11:E11)
    "start_time": "F11",         # 開始 (merged F11:G11)
    "end_time": "I11",           # 終了 (merged I11:J11)
    "location": "C12",           # 作業場所 (merged C12:F12)
    "people_count": "G12",       # 作業人数 (merged G12:I12)
    "work_content_1": "C14",     # 作業内容 (merged C14:J14)
    "work_content_2": "C15",     # 作業内容 2行目 (merged C15:J15)
    "focus_instructions": "B25", # 重点指示事項 (merged B25:J27)
    "notes": "B48",              # 連絡事項 (merged B48:J50)
}

# Check items (prefix ✓ into the item text cell)
HAZARD_ITEMS = {
    "感電・漏電事故": "B17",
    "火災": "C17",
    "停電事故": "D17",
    "漏電事故": "F17",
    "墜落・落下事故": "B18",
    "酸欠事故": "C18",
    "騒音、振動、異臭、埃等のクレーム": "D18",
    "その他(危険ポイント)": "B19",  # text includes ⑧その他（   ）, we will inject detail
}
AVOID_ITEMS = {
    "活線作業の禁止": "B21",
    "不良工具の使用禁止": "C21",
    "保護具使用": "E21",
    "ヘルメット着用": "B22",
    "安全帯着用": "C22",
    "安全柵取付": "E22",
    "消火器設置": "G22",
    "作業時間帯調整": "B23",
    "その他(危険回避)": "C23",
}
FINISH_ITEMS = {
    "電源・スイッチ・バルブ等の復旧": "B44",
    "火気・危険物作業実施後の安全確認": "E44",
    "不要品の搬出及び清掃": "B45",
    "借用品の返却": "E45",
    "部屋の施錠": "B46",
    "その他(終了確認)": "E46",
}

def _connect():
    # Ensure the parent directory exists (Render Free: use /tmp by default)
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    con = sqlite3.connect(DB_PATH, check_same_thread=False)
    con.row_factory = sqlite3.Row
    return con

def _init_db():
    con = _connect()
    cur = con.cursor()
    cur.executescript("""
    CREATE TABLE IF NOT EXISTS companies (
      company_id TEXT PRIMARY KEY,
      company_name TEXT NOT NULL,
      password_hash BLOB NOT NULL,
      is_admin INTEGER NOT NULL DEFAULT 0,
      is_enabled INTEGER NOT NULL DEFAULT 1,
      created_at TEXT NOT NULL
    );

    CREATE TABLE IF NOT EXISTS name_candidates (
      company_id TEXT NOT NULL,
      name TEXT NOT NULL,
      PRIMARY KEY (company_id, name),
      FOREIGN KEY (company_id) REFERENCES companies(company_id)
    );

    CREATE TABLE IF NOT EXISTS ky_records (
      id TEXT PRIMARY KEY,
      company_id TEXT NOT NULL,
      created_at TEXT NOT NULL,
      updated_at TEXT NOT NULL,
      inputter_name TEXT NOT NULL,
      work_title TEXT,
      work_company TEXT,
      phone TEXT,
      work_date TEXT,
      start_time TEXT,
      end_time TEXT,
      location TEXT,
      people_count TEXT,
      work_content TEXT,
      hazards_json TEXT,
      hazards_other TEXT,
      avoid_json TEXT,
      avoid_other TEXT,
      focus_instructions TEXT,
      finish_json TEXT,
      finish_other TEXT,
      notes TEXT,
      FOREIGN KEY (company_id) REFERENCES companies(company_id)
    );
    """)
    con.commit()
    con.close()

def _seed_if_needed():
    if not os.path.exists(SEED_PATH):
        return
    con = _connect()
    cur = con.cursor()
    cur.execute("SELECT COUNT(*) AS n FROM companies")
    if cur.fetchone()["n"] > 0:
        con.close()
        return

    with open(SEED_PATH, "r", encoding="utf-8") as f:
        seed = json.load(f)

    now = datetime.utcnow().isoformat()
    # credentials from seed.json (plaintext) are hashed here and never stored as plaintext in DB
    cred_map = {c["company_id"]: c["password"] for c in seed.get("initial_credentials", [])}

    for c in seed["companies"]:
        cid = c["company_id"]
        cname = c["company_name"]
        is_admin = 1 if c.get("is_admin") else 0
        pw = cred_map.get(cid, "ChangeMe123!")
        pw_hash = bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt())
        cur.execute(
            "INSERT INTO companies(company_id, company_name, password_hash, is_admin, is_enabled, created_at) VALUES (?,?,?,?,?,?)",
            (cid, cname, pw_hash, is_admin, 1, now),
        )

    for cid, names in seed.get("name_candidates", {}).items():
        for nm in names:
            cur.execute("INSERT INTO name_candidates(company_id, name) VALUES (?,?)", (cid, nm))

    con.commit()
    con.close()

def _apply_retention():
    # Delete records older than RETENTION_YEARS
    cutoff = datetime.utcnow() - relativedelta(years=RETENTION_YEARS)
    con = _connect()
    cur = con.cursor()
    cur.execute("DELETE FROM ky_records WHERE created_at < ?", (cutoff.isoformat(),))
    con.commit()
    con.close()

def _verify_login(company_id: str, password: str):
    con = _connect()
    cur = con.cursor()
    cur.execute("SELECT * FROM companies WHERE company_id=?", (company_id,))
    row = cur.fetchone()
    con.close()
    if not row:
        return None, "IDが見つかりません。"
    if row["is_enabled"] != 1:
        return None, "このアカウントは停止されています。管理者に連絡してください。"
    if not bcrypt.checkpw(password.encode("utf-8"), row["password_hash"]):
        return None, "パスワードが違います。"
    return dict(row), None

def _list_candidates(company_id: str):
    con = _connect()
    cur = con.cursor()
    cur.execute("SELECT name FROM name_candidates WHERE company_id=? ORDER BY name", (company_id,))
    rows = [r["name"] for r in cur.fetchall()]
    con.close()
    return rows

def _add_candidate(company_id: str, name: str):
    con = _connect()
    cur = con.cursor()
    cur.execute("INSERT OR IGNORE INTO name_candidates(company_id, name) VALUES (?,?)", (company_id, name))
    con.commit()
    con.close()

def _new_id():
    # simple unique id
    import secrets
    return secrets.token_hex(12)

def _save_record(data: dict, record_id: str | None = None):
    now = datetime.utcnow().isoformat()
    con = _connect()
    cur = con.cursor()
    if record_id is None:
        record_id = _new_id()
        cur.execute("""
        INSERT INTO ky_records(
          id, company_id, created_at, updated_at, inputter_name,
          work_title, work_company, phone, work_date, start_time, end_time, location, people_count, work_content,
          hazards_json, hazards_other, avoid_json, avoid_other, focus_instructions,
          finish_json, finish_other, notes
        ) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        """, (
            record_id, data["company_id"], now, now, data["inputter_name"],
            data.get("work_title",""), data.get("work_company",""), data.get("phone",""),
            data.get("work_date",""), data.get("start_time",""), data.get("end_time",""),
            data.get("location",""), data.get("people_count",""), data.get("work_content",""),
            json.dumps(data.get("hazards",[]), ensure_ascii=False),
            data.get("hazards_other",""),
            json.dumps(data.get("avoid",[]), ensure_ascii=False),
            data.get("avoid_other",""),
            data.get("focus_instructions",""),
            json.dumps(data.get("finish",[]), ensure_ascii=False),
            data.get("finish_other",""),
            data.get("notes",""),
        ))
    else:
        cur.execute("""
        UPDATE ky_records SET
          updated_at=?, inputter_name=?,
          work_title=?, work_company=?, phone=?, work_date=?, start_time=?, end_time=?, location=?, people_count=?, work_content=?,
          hazards_json=?, hazards_other=?, avoid_json=?, avoid_other=?, focus_instructions=?,
          finish_json=?, finish_other=?, notes=?
        WHERE id=? AND company_id=?
        """, (
            now, data["inputter_name"],
            data.get("work_title",""), data.get("work_company",""), data.get("phone",""),
            data.get("work_date",""), data.get("start_time",""), data.get("end_time",""),
            data.get("location",""), data.get("people_count",""), data.get("work_content",""),
            json.dumps(data.get("hazards",[]), ensure_ascii=False),
            data.get("hazards_other",""),
            json.dumps(data.get("avoid",[]), ensure_ascii=False),
            data.get("avoid_other",""),
            data.get("focus_instructions",""),
            json.dumps(data.get("finish",[]), ensure_ascii=False),
            data.get("finish_other",""),
            data.get("notes",""),
            record_id, data["company_id"]
        ))
    con.commit()
    con.close()
    return record_id

def _load_records(company_id: str, limit: int = 50):
    con = _connect()
    cur = con.cursor()
    cur.execute("""
    SELECT id, created_at, updated_at, inputter_name, work_title, work_date, location
    FROM ky_records
    WHERE company_id=?
    ORDER BY created_at DESC
    LIMIT ?
    """, (company_id, limit))
    rows = [dict(r) for r in cur.fetchall()]
    con.close()
    return rows

def _load_record(company_id: str, record_id: str):
    con = _connect()
    cur = con.cursor()
    cur.execute("SELECT * FROM ky_records WHERE company_id=? AND id=?", (company_id, record_id))
    row = cur.fetchone()
    con.close()
    if not row:
        return None
    d = dict(row)
    d["hazards"] = json.loads(d.get("hazards_json") or "[]")
    d["avoid"] = json.loads(d.get("avoid_json") or "[]")
    d["finish"] = json.loads(d.get("finish_json") or "[]")
    return d

def _admin_list_companies():
    con = _connect()
    cur = con.cursor()
    cur.execute("SELECT company_id, company_name, is_enabled, is_admin, created_at FROM companies ORDER BY is_admin DESC, company_name")
    rows = [dict(r) for r in cur.fetchall()]
    con.close()
    return rows

def _admin_set_enabled(company_id: str, enabled: bool):
    con = _connect()
    cur = con.cursor()
    cur.execute("UPDATE companies SET is_enabled=? WHERE company_id=? AND is_admin=0", (1 if enabled else 0, company_id))
    con.commit()
    con.close()

def _admin_reset_password(company_id: str):
    import secrets, string
    alphabet = string.ascii_letters + string.digits
    new_pw = ''.join(secrets.choice(alphabet) for _ in range(18))
    pw_hash = bcrypt.hashpw(new_pw.encode("utf-8"), bcrypt.gensalt())
    con = _connect()
    cur = con.cursor()
    cur.execute("UPDATE companies SET password_hash=? WHERE company_id=? AND is_admin=0", (pw_hash, company_id))
    con.commit()
    con.close()
    return new_pw

def _prefix_check(text: str, checked: bool):
    if checked:
        if text.startswith("✓"):
            return text
        return "✓" + text
    else:
        # remove leading check if present
        return text[1:] if text.startswith("✓") else text

def _inject_other(base_text: str, detail: str):
    # replace inside '（ ... ）' with detail (keep parentheses)
    if "（" in base_text and "）" in base_text:
        pre = base_text.split("（",1)[0]
        post = base_text.split("）",1)[1]
        return f"{pre}（{detail}）{post}"
    return base_text + f"（{detail}）"

def _render_excel(record: dict) -> bytes:
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb["安全指示ＫＹ記録書"]

    # Fill basics
    ws[CELL["work_title"]] = record.get("work_title","")
    ws[CELL["work_company"]] = record.get("work_company","")
    ws[CELL["phone"]] = record.get("phone","")
    ws[CELL["work_date"]] = record.get("work_date","")
    ws[CELL["start_time"]] = record.get("start_time","")
    ws[CELL["end_time"]] = record.get("end_time","")
    ws[CELL["location"]] = record.get("location","")
    ws[CELL["people_count"]] = record.get("people_count","")

    # Work content split into two lines (template has two merged rows)
    content = (record.get("work_content") or "").strip()
    lines = content.splitlines()
    if len(lines) == 0:
        ws[CELL["work_content_1"]] = ""
        ws[CELL["work_content_2"]] = ""
    elif len(lines) == 1:
        ws[CELL["work_content_1"]] = lines[0]
        ws[CELL["work_content_2"]] = ""
    else:
        ws[CELL["work_content_1"]] = lines[0]
        ws[CELL["work_content_2"]] = "\n".join(lines[1:])

    # Focus instructions
    focus = record.get("focus_instructions","").strip()
    # always include inputter name for audit
    inputter = record.get("inputter_name","").strip()
    if inputter:
        focus = (focus + "\n" if focus else "") + f"【入力者】{inputter}"
    ws[CELL["focus_instructions"]] = focus

    # Notes
    ws[CELL["notes"]] = (record.get("notes") or "").strip()

    # Apply check items by prefixing ✓
    hazards_selected = set(record.get("hazards") or [])
    hazards_other = (record.get("hazards_other") or "").strip()

    for label, cell in HAZARD_ITEMS.items():
        base = ws[cell].value or ""
        if label.startswith("その他"):
            if hazards_other:
                ws[cell] = _prefix_check(_inject_other(base, hazards_other), True)
            else:
                ws[cell] = _prefix_check(base, False)
        else:
            ws[cell] = _prefix_check(base, label in hazards_selected)

    avoid_selected = set(record.get("avoid") or [])
    avoid_other = (record.get("avoid_other") or "").strip()
    for label, cell in AVOID_ITEMS.items():
        base = ws[cell].value or ""
        if label.startswith("その他"):
            if avoid_other:
                ws[cell] = _prefix_check(_inject_other(base, avoid_other), True)
            else:
                ws[cell] = _prefix_check(base, False)
        else:
            ws[cell] = _prefix_check(base, label in avoid_selected)

    finish_selected = set(record.get("finish") or [])
    finish_other = (record.get("finish_other") or "").strip()
    for label, cell in FINISH_ITEMS.items():
        base = ws[cell].value or ""
        if label.startswith("その他"):
            if finish_other:
                ws[cell] = _prefix_check(_inject_other(base, finish_other), True)
            else:
                ws[cell] = _prefix_check(base, False)
        else:
            ws[cell] = _prefix_check(base, label in finish_selected)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def _login_view():
    st.subheader("ログイン")
    with st.form("login"):
        company_id = st.text_input("会社ID", placeholder="例）trust-bosai", autocomplete="username")
        password = st.text_input("パスワード", type="password", autocomplete="current-password")
        ok = st.form_submit_button("ログイン")
    if ok:
        user, err = _verify_login(company_id.strip(), password)
        if err:
            st.error(err)
            return
        st.session_state["auth"] = user
        st.success("ログインしました。")
        st.rerun()

def _logout_button():
    if st.button("ログアウト"):
        st.session_state.pop("auth", None)
        st.rerun()

def _admin_panel():
    st.divider()
    st.subheader("管理者メニュー")
    st.caption("会社アカウントの停止/再開、パスワード再発行ができます。")

    companies = _admin_list_companies()
    for c in companies:
        if c["is_admin"] == 1:
            continue
        col1, col2, col3, col4 = st.columns([3,2,2,3])
        with col1:
            st.write(f"**{c['company_name']}**  (`{c['company_id']}`)")
            st.caption("有効" if c["is_enabled"]==1 else "停止中")
        with col2:
            if c["is_enabled"]==1:
                if st.button("停止", key=f"disable_{c['company_id']}"):
                    _admin_set_enabled(c["company_id"], False)
                    st.rerun()
            else:
                if st.button("再開", key=f"enable_{c['company_id']}"):
                    _admin_set_enabled(c["company_id"], True)
                    st.rerun()
        with col3:
            if st.button("PW再発行", key=f"reset_{c['company_id']}"):
                new_pw = _admin_reset_password(c["company_id"])
                st.session_state[f"newpw_{c['company_id']}"] = new_pw
        with col4:
            new_pw = st.session_state.get(f"newpw_{c['company_id']}")
            if new_pw:
                st.warning(f"新PW（1回表示）: `{new_pw}`")
                st.caption("必ず控えてから画面更新してください。")
    st.divider()

def _record_form(default: dict | None, company_id: str):
    candidates = _list_candidates(company_id)

    st.subheader("KY入力（保存→Excel出力）")
    st.caption("※入力者名は必須です（会社共通ID運用のため）。")

    inputter = st.selectbox("入力者名（必須）", options=([""] + candidates + ["（手入力）"]), index=0)
    inputter_custom = ""
    if inputter == "（手入力）":
        inputter_custom = st.text_input("入力者名を入力", value=(default.get("inputter_name","") if default else ""))
        inputter_name = inputter_custom.strip()
    else:
        inputter_name = inputter.strip() if inputter else (default.get("inputter_name","").strip() if default else "")

    col1, col2 = st.columns(2)
    with col1:
        work_title = st.text_input("作業件名", value=(default.get("work_title","") if default else ""))
        work_company = st.text_input("作業会社名", value=(default.get("work_company","") if default else ""))
        phone = st.text_input("電話番号", value=(default.get("phone","") if default else ""))
        work_date = st.text_input("作業予定日（例：2026/02/19）", value=(default.get("work_date","") if default else ""))
    with col2:
        start_time = st.text_input("開始（例：01:00）", value=(default.get("start_time","") if default else ""))
        end_time = st.text_input("終了（例：07:00）", value=(default.get("end_time","") if default else ""))
        location = st.text_input("作業場所", value=(default.get("location","") if default else ""))
        people_count = st.text_input("作業人数", value=(default.get("people_count","") if default else ""))

    work_content = st.text_area("作業内容（改行OK）", height=120, value=(default.get("work_content","") if default else ""))

    st.markdown("### 想定される危険ポイント")
    hazards = st.multiselect("該当するものにチェック", options=[
        "感電・漏電事故","火災","停電事故","漏電事故","墜落・落下事故","酸欠事故","騒音、振動、異臭、埃等のクレーム"
    ], default=(default.get("hazards",[]) if default else []))
    hazards_other = st.text_input("その他（危険ポイント）", value=(default.get("hazards_other","") if default else ""), placeholder="例：挟まれ、切創 など")

    st.markdown("### 危険回避のポイント")
    avoid = st.multiselect("該当するものにチェック", options=[
        "活線作業の禁止","不良工具の使用禁止","保護具使用","ヘルメット着用","安全帯着用","安全柵取付","消火器設置","作業時間帯調整"
    ], default=(default.get("avoid",[]) if default else []))
    avoid_other = st.text_input("その他（危険回避）", value=(default.get("avoid_other","") if default else ""), placeholder="例：立入禁止・誘導員配置 など")

    focus_instructions = st.text_area("施設管理担当者からの重点指示事項（必要なら）", height=80, value=(default.get("focus_instructions","") if default else ""))

    st.markdown("### 作業終了確認")
    finish = st.multiselect("該当するものにチェック", options=[
        "電源・スイッチ・バルブ等の復旧","火気・危険物作業実施後の安全確認","不要品の搬出及び清掃","借用品の返却","部屋の施錠"
    ], default=(default.get("finish",[]) if default else []))
    finish_other = st.text_input("その他（終了確認）", value=(default.get("finish_other","") if default else ""))

    notes = st.text_area("連絡事項（任意）", height=80, value=(default.get("notes","") if default else ""))

    payload = {
        "company_id": company_id,
        "inputter_name": inputter_name,
        "work_title": work_title,
        "work_company": work_company,
        "phone": phone,
        "work_date": work_date,
        "start_time": start_time,
        "end_time": end_time,
        "location": location,
        "people_count": people_count,
        "work_content": work_content,
        "hazards": hazards,
        "hazards_other": hazards_other,
        "avoid": avoid,
        "avoid_other": avoid_other,
        "focus_instructions": focus_instructions,
        "finish": finish,
        "finish_other": finish_other,
        "notes": notes,
    }

    return payload

def main():
    st.set_page_config(page_title=APP_TITLE, layout="centered")
    st.title(APP_TITLE)
    st.caption("会社別ログイン／自社履歴閲覧可／保存3年／Excel書式固定出力")

    _init_db()
    _seed_if_needed()
    _apply_retention()

    auth = st.session_state.get("auth")
    if not auth:
        _login_view()
        st.info("※会社ID/初期パスワードは管理者から共有されます。")
        return

    # header
    c1, c2 = st.columns([4,1])
    with c1:
        st.write(f"ログイン中：**{auth['company_name']}** (`{auth['company_id']}`)")
    with c2:
        _logout_button()

    if auth.get("is_admin") == 1:
        st.success("管理者モードです（全社管理が可能）。")

    # Tabs
    tab1, tab2 = st.tabs(["新規作成 / 編集", "自社履歴（複製）"])

    # Self records
    with tab2:
        st.subheader("自社履歴")
        st.caption("自社で作成したKYだけ表示されます。クリックで複製・再編集できます。")
        records = _load_records(auth["company_id"], limit=80)
        if not records:
            st.info("まだ履歴がありません。")
        else:
            # simple list
            for r in records:
                title = r.get("work_title") or "(無題)"
                created = r["created_at"][:19].replace("T"," ")
                label = f"{created}｜{title}｜{r.get('location','')}"
                if st.button(label, key=f"pick_{r['id']}"):
                    st.session_state["editing_id"] = r["id"]
                    st.rerun()

    # Create/edit
    with tab1:
        edit_id = st.session_state.get("editing_id")
        default = None
        if edit_id:
            default = _load_record(auth["company_id"], edit_id)
            if not default:
                st.warning("選択した履歴が見つかりませんでした。")
                st.session_state.pop("editing_id", None)
                st.rerun()

        if edit_id:
            st.subheader("編集 / 再出力")
            cols = st.columns([1,1,2])
            with cols[0]:
                if st.button("新規作成に戻る"):
                    st.session_state.pop("editing_id", None)
                    st.rerun()
            with cols[1]:
                if st.button("この内容を複製して新規"):
                    st.session_state["clone_from"] = edit_id
                    st.session_state.pop("editing_id", None)
                    st.rerun()

        clone_from = st.session_state.get("clone_from")
        if clone_from and not edit_id:
            base = _load_record(auth["company_id"], clone_from)
            default = base
            st.session_state.pop("clone_from", None)
            st.info("複製元の内容を読み込みました。必要箇所だけ修正して保存してください。")

        payload = _record_form(default, auth["company_id"])

        # Validate inputter
        if not payload["inputter_name"]:
            st.warning("入力者名は必須です。")

        btn_col1, btn_col2 = st.columns([1,1])
        with btn_col1:
            if st.button("保存"):
                if not payload["inputter_name"]:
                    st.error("入力者名が未入力です。")
                else:
                    # if custom input and not in candidates, add it for next time
                    _add_candidate(auth["company_id"], payload["inputter_name"])
                    saved_id = _save_record(payload, record_id=edit_id if edit_id else None)
                    st.session_state["editing_id"] = saved_id
                    st.success("保存しました。")
                    st.rerun()

        with btn_col2:
            if st.button("保存してExcel出力"):
                if not payload["inputter_name"]:
                    st.error("入力者名が未入力です。")
                else:
                    _add_candidate(auth["company_id"], payload["inputter_name"])
                    saved_id = _save_record(payload, record_id=edit_id if edit_id else None)
                    rec = _load_record(auth["company_id"], saved_id)
                    xbytes = _render_excel(rec)
                    filename = f"KY_{auth['company_id']}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
                    st.download_button("Excelをダウンロード", data=xbytes, file_name=filename,
                                       mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
                    st.session_state["editing_id"] = saved_id
                    st.success("保存＆Excel生成しました（上のボタンからDL）。")

    # Admin
    if auth.get("is_admin") == 1:
        _admin_panel()

if __name__ == "__main__":
    main()
